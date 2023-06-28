train_transform = pd.DataFrame(train_set.drop(['Lm0', 'Hm0','q','qAD'], axis=1))
desc_train_transform = train_transform.describe()
desc_train_transform.to_excel('desc_train_transform.xlsx') 
train_transform = pd.DataFrame(scaler.fit_transform(train_transform))
train_transform.columns = train_set.drop(['Lm0','Hm0','q','qAD'], axis=1).columns

test_transform = pd.DataFrame(test_set.drop(['Lm0','Hm0','q','qAD'], axis=1))

#save description of test data
desc_test_transform = test_transform.describe()
desc_test_transform.to_csv('desc_test_transform.csv', decimal=".", sep=' ', header = True, index = False ) 

#add noise i zapisz ponownie
#jesli dla zaszumionego datasetu będzie podobny wynik albo mniejszy - to jest ok, a jesli wiekszy to overfitting i 
test_transform = pd.DataFrame(scaler.transform(test_transform))
test_transform.columns = test_set.drop(['Lm0','Hm0','q','qAD'], axis=1).columns

fin_ANN_1_qad_test_transform = pd.DataFrame(scaler.fit_transform(ANN_1_qad.values.reshape(-1,1)))

fin_nn_overtoppingANN_1_transform_qad = pd.DataFrame(scaler.transform(nn_overtoppingANN_1_transform_qad.values.reshape(-1,1)))
print("############## SCORE DATASET TEST FOR rf ########")
print()
print("R2 TEST")
print(r2_score(fin_ANN_1_qad_test_transform, fin_nn_overtoppingANN_1_transform_qad))
print()
print("RMSE TEST")
print(np.sqrt(mean_squared_error(fin_ANN_1_qad_test_transform, fin_nn_overtoppingANN_1_transform_qad)))
rf = RandomForestRegressor(     max_depth=25,
                                n_estimators=bst_n_estimators,
                                random_state=random_state)
rf.fit(train['features'], train['target'].ravel())
prediction = rf.predict(test['features'])
test['predictions'] = pd.DataFrame(prediction)

print("############## SCORE DATASET TEST FOR rf ########")
print()
print("R2 TEST")
print(r2_score(test['target'], test['predictions']))
print()
print("RMSE TEST")
print(np.sqrt(mean_squared_error(test['target'], test['predictions'])))
print("############## SCORE DATASET TEST FOR rf ########")
print()
print("R2 TEST")
print(r2_score(test['target'], test['predictions']))
print()
print("RMSE TEST")
print(np.sqrt(mean_squared_error(test['target'], test['predictions'])))
def RandomForestR(params,  scoring, features, target):
    
    rf = RandomForestRegressor(random_state)
    
    for x in scoring:
        grid_search = GridSearchCV(estimator = rf, param_grid = params, 
                                   n_jobs = -1, verbose = 2, scoring=x)
        
        grid_search.fit(features, target)
        print()
        print("######### Tuning hyper-parameters for %s" % x)
        print()
        print()
        print("Best parameters set found on development set:")
        if x == 'neg_mean_squared_error':
            print(grid_search.best_params_, ' -----> SCORE RMSE: %f' %
                  np.sqrt(-grid_search.best_score_))
        else:
            print(grid_search.best_params_, ' -----> SCORE: %f' %
                  grid_search.best_score_)
# search hyperparameters

scoring = ['neg_mean_squared_error', 'r2']
rf = RandomForestRegressor(     max_depth=25,
                                n_estimators=195,
                                random_state=random_state)
rf.fit(train['features'], train['target'].ravel())
prediction = rf.predict(test['features'])
test['predictions'] = pd.DataFrame(prediction)
print("############## SCORE DATASET TEST FOR rf ########")
print()
print("R2 TEST")
print(r2_score(test['target'], test['predictions']))
print()
print("RMSE TEST")
print(np.sqrt(mean_squared_error(test['target'], test['predictions']))) 



print("############## SCORE DATASET TEST FOR rf ########")
print()
print("R2 TEST")
print(r2_score(fin_ANN_1_qad_test_transform, fin_nn_overtoppingANN_1_transform_qad))
print()
print("RMSE TEST")
print(np.sqrt(mean_squared_error(fin_ANN_1_qad_test_transform, fin_nn_overtoppingANN_1_transform_qad)))
@author: Dorota Orzechowska
"""
# Import libraries

import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler,MinMaxScaler
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import learning_curve
from sklearn.model_selection import cross_val_score
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error


from sklearn.ensemble import RandomForestRegressor

import seaborn as sns

import warnings
warnings.filterwarnings('ignore')

random_state = 134

# Load the data 
df= pd.read_excel('Database_Eurotop_v5.xlsx')

#dfx= pd.read_excel('Database_Eurotop_v5.xlsx')
df.head()

#create delete list for parameters unused
delete_list = ['Label']

#delete non-core data
df.drop(df.index[df['Core data'] != 'Z'], inplace=True)

#delete rows with Nan values of'q'
df = df.dropna(subset=['q'])

# delete simulation with RF or CF = 4 or 3
df.drop(df.index[df['RF'] == 4], inplace=True)
df.drop(df.index[df['CF'] == 4], inplace=True)
df.drop(df.index[df['RF'] == 3], inplace=True)
df.drop(df.index[df['CF'] == 3], inplace=True)

dfsave = pd.DataFrame({
    'Label':df['Label'],
    'm':df['mmm'],
    'h':df['h'],
    'Hm0 toe':df['Hm0 toe'],
    'Tm1,0t':df['Tm1,0t'],
    'β':df['β'],
    'ht':df['ht'],
    'Bt':df['Bt'],
    'hb':df['hb'],
    'B':df['B'],
    'cotad':df['cotαd'],
    'cotaincl':df['cotαincl'],
    'cotau':df['cotαu'],
    'gf_d':df['gf_d'],
    'gf_u':df['gf_u'],
    'gf':df['gf'],
    'D':df['D'],
    'D50_d':df['D50_d'],
    'D50_u':df['D50_u'],
    'Ac':df['Ac'],
    'Rc':df['Rc'],    
    'Gc':df['Gc'],
    'Kr':df['Kr'],
    'Kt':df['Kt'],
    'q' :df['q']
    })



#delete 'Label' to change data type for float
for x in delete_list:
    dfsave = dfsave.drop(columns=[x])

# convert in float
#dfsave.drop(dfsave.head(1).index, inplace=True)
dfsave = dfsave.astype('float')
dfsave = dfsave.dropna(subset=['m'])
dfsave = dfsave.dropna(subset=['D'])
a = dfsave.describe()



#delete q == 0
dfsave.drop(dfsave.index[dfsave['q'] == 0 ], inplace=True)


#%matplotlib qt

for i in dfsave.columns:
    fig = plt.hist(dfsave[i],bins=10)
    plt.title(i)
    plt.show(fig)

# delete the NaN 
#dfsave = dfsave.dropna()
#a = df.describe()

#corr matrix to see correlation of parameters
corr_matrix = dfsave.corr()
f, ax = plt.subplots(figsize=(20,20))
g = sns.heatmap(corr_matrix, cmap = sns.diverging_palette(10, 250, as_cmap = True), annot = True)

#creating an additional feature - compute spectral wave length in deep water lm0 in meter
dfsave['Lm0'] = ((9.81*(dfsave['Tm1,0t'])**2)/(2*np.pi))

# create dataframe containing the model parameters
dff = pd.DataFrame({
    'Lm0' :dfsave['Lm0'],
    'Hm0' :dfsave['Hm0 toe'],
    'Hm0/Lm0':dfsave['Hm0 toe']/dfsave['Lm0'],
    'β':dfsave['β'],
    'h/Lm0':dfsave['h']/dfsave['Lm0'],
    'ht/Hm0':dfsave['ht']/dfsave['Hm0 toe'],
    'Bt/Lm0':dfsave['Bt']/dfsave['Lm0'],
    'hb/Hm0':dfsave['hb']/dfsave['Hm0 toe'],
    'B/Lm0':dfsave['B']/dfsave['Lm0'],
    'Ac/Hm0':dfsave['Ac']/dfsave['Hm0 toe'],
    'Rc/Hm0':dfsave['Rc']/dfsave['Hm0 toe'],
    #'Rc':dfsave['Rc'],
    'Gc/Lm0':dfsave['Gc']/dfsave['Lm0'],
    'm':dfsave['m'],
    'cotad':dfsave['cotad'],
    'cotaincl':dfsave['cotaincl'],
    'gf':dfsave['gf'],
    'q':dfsave['q'],
    'D/Hm0':dfsave['D']/dfsave['Hm0 toe'],
    #target:
    'qAD':np.log10(dfsave['q']/np.sqrt(9.81*(dfsave['Hm0 toe'])**3)),
    })
#dff.head()

dff.describe()

#dff = dff.dropna()

b= dff.describe()

"""
Created on Sat Dec 31 19:49:44 2022

@author: Dorota Orzechowska
"""
# Import libraries

import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler,MinMaxScaler
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import learning_curve
from sklearn.model_selection import cross_val_score
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error


from sklearn.ensemble import RandomForestRegressor

import seaborn as sns

import warnings
warnings.filterwarnings('ignore')

random_state = 134

# Load the data 
df= pd.read_excel('Database_Eurotop_v5.xlsx')

#dfx= pd.read_excel('Database_Eurotop_v5.xlsx')
df.head()

#create delete list for parameters unused
delete_list = ['Label']

#delete non-core data
df.drop(df.index[df['Core data'] != 'Z'], inplace=True)

#delete rows with Nan values of'q'
df = df.dropna(subset=['q'])

# delete simulation with RF or CF = 4 or 3
df.drop(df.index[df['RF'] == 4], inplace=True)
df.drop(df.index[df['CF'] == 4], inplace=True)
df.drop(df.index[df['RF'] == 3], inplace=True)
df.drop(df.index[df['CF'] == 3], inplace=True)

dfsave = pd.DataFrame({
    'Label':df['Label'],
    'm':df['mmm'],
    'h':df['h'],
    'Hm0 toe':df['Hm0 toe'],
    'Tm1,0t':df['Tm1,0t'],
    'β':df['β'],
    'ht':df['ht'],
    'Bt':df['Bt'],
    'hb':df['hb'],
    'B':df['B'],
    'cotad':df['cotαd'],
    'cotaincl':df['cotαincl'],
    'cotau':df['cotαu'],
    'gf_d':df['gf_d'],
    'gf_u':df['gf_u'],
    'gf':df['gf'],
    'D':df['D'],
    'D50_d':df['D50_d'],
    'D50_u':df['D50_u'],
    'Ac':df['Ac'],
    'Rc':df['Rc'],    
    'Gc':df['Gc'],
    'Kr':df['Kr'],
    'Kt':df['Kt'],
    'q' :df['q']
    })



#delete 'Label' to change data type for float
for x in delete_list:
    dfsave = dfsave.drop(columns=[x])

# convert in float
#dfsave.drop(dfsave.head(1).index, inplace=True)
dfsave = dfsave.astype('float')
dfsave = dfsave.dropna(subset=['m'])
dfsave = dfsave.dropna(subset=['D'])
a = dfsave.describe()



#delete q == 0
dfsave.drop(dfsave.index[dfsave['q'] == 0 ], inplace=True)


#%matplotlib qt

for i in dfsave.columns:
    fig = plt.hist(dfsave[i],bins=10)
    plt.title(i)
    plt.show(fig)

# delete the NaN 
#dfsave = dfsave.dropna()
#a = df.describe()

#corr matrix to see correlation of parameters
corr_matrix = dfsave.corr()
f, ax = plt.subplots(figsize=(20,20))
g = sns.heatmap(corr_matrix, cmap = sns.diverging_palette(10, 250, as_cmap = True), annot = True)

#creating an additional feature - compute spectral wave length in deep water lm0 in meter
dfsave['Lm0'] = ((9.81*(dfsave['Tm1,0t'])**2)/(2*np.pi))

# create dataframe containing the model parameters
dff = pd.DataFrame({
    'Lm0' :dfsave['Lm0'],
    'Hm0' :dfsave['Hm0 toe'],
    'Hm0/Lm0':dfsave['Hm0 toe']/dfsave['Lm0'],
    'β':dfsave['β'],
    'h/Lm0':dfsave['h']/dfsave['Lm0'],
    'ht/Hm0':dfsave['ht']/dfsave['Hm0 toe'],
    'Bt/Lm0':dfsave['Bt']/dfsave['Lm0'],
    'hb/Hm0':dfsave['hb']/dfsave['Hm0 toe'],
    'B/Lm0':dfsave['B']/dfsave['Lm0'],
    'Ac/Hm0':dfsave['Ac']/dfsave['Hm0 toe'],
    'Rc/Hm0':dfsave['Rc']/dfsave['Hm0 toe'],
    #'Rc':dfsave['Rc'],
    'Gc/Lm0':dfsave['Gc']/dfsave['Lm0'],
    'm':dfsave['m'],
    'cotad':dfsave['cotad'],
    'cotaincl':dfsave['cotaincl'],
    'gf':dfsave['gf'],
    'q':dfsave['q'],
    'D/Hm0':dfsave['D']/dfsave['Hm0 toe'],
    #target:
    'qAD':np.log10(dfsave['q']/np.sqrt(9.81*(dfsave['Hm0 toe'])**3)),
    })
#dff.head()

dff.describe()

#dff = dff.dropna()

b= dff.describe()

# Replace infinite updated data with nan
dff.replace([np.inf, -np.inf], np.nan, inplace=True)
# Drop rows with NaN
dff.dropna(inplace=True)

#dfsave.to_csv('NN_OVERTOPPING6.csv', decimal=".", sep=' ', header = False, index = False ) 
#val_=dfsave.describe()

corr_matrix = dff.corr()
f, ax = plt.subplots(figsize=(20,20))
g = sns.heatmap(corr_matrix, cmap = sns.diverging_palette(10, 250, as_cmap = True), annot = True)


train_set, test_set = train_test_split(dff, test_size=0.2, random_state = random_state)

#scaler_std = StandardScaler()
scaler = StandardScaler()

ANN_1 = dfsave.loc[test_set.index] 
ANN_1_q = pd.DataFrame(ANN_1['q'])
ANN_1_qad = pd.DataFrame(np.log10(ANN_1['q']/np.sqrt(9.81*(ANN_1['Hm0 toe'])**3)))

ANN_1 = ANN_1.drop(['Lm0', 'D', 'gf','Kr', 'Kt','q', 'cotaincl'], axis = 1)



#prepare file for ANN tool
ANN_1['Kr'] = 0
ANN_1['Kt'] = 0
ANN_1['q'] = 1
ANN_1.insert(loc=0, column='Label', value=df['Label'])
ANN_1.to_csv('ANN_1.csv', decimal=".", sep='|', index = False ) 

nn_overtoppingANN_1 = pd.read_csv('ANN_results.csv', decimal=".", sep='|')

nn_overtoppingANN_1_transform = pd.DataFrame(nn_overtoppingANN_1['Average'])

ANN_1_Hm = pd.DataFrame(ANN_1["Hm0 toe"])
ANN_1_Hm=ANN_1_Hm.reset_index()
ANN_1_Hm = ANN_1_Hm.drop(['index'], axis=1)

nn_overtoppingANN_1_transform_qad = pd.DataFrame(np.log10(nn_overtoppingANN_1_transform['Average']/np.sqrt(9.81*(ANN_1_Hm['Hm0 toe'])**3)))


# save description of train data
train_transform = pd.DataFrame(train_set.drop(['Lm0', 'Hm0','q','qAD'], axis=1))
desc_train_transform = train_transform.describe()
desc_train_transform.to_excel('desc_train_transform.xlsx') 
train_transform = pd.DataFrame(scaler.fit_transform(train_transform))
train_transform.columns = train_set.drop(['Lm0','Hm0','q','qAD'], axis=1).columns

test_transform = pd.DataFrame(test_set.drop(['Lm0','Hm0','q','qAD'], axis=1))

#save description of test data
desc_test_transform = test_transform.describe()
desc_test_transform.to_csv('desc_test_transform.csv', decimal=".", sep=' ', header = True, index = False ) 

#add noise i zapisz ponownie
#jesli dla zaszumionego datasetu będzie podobny wynik albo mniejszy - to jest ok, a jesli wiekszy to overfitting i 
test_transform = pd.DataFrame(scaler.transform(test_transform))
test_transform.columns = test_set.drop(['Lm0','Hm0','q','qAD'], axis=1).columns

fin_ANN_1_qad_test_transform = pd.DataFrame(scaler.fit_transform(ANN_1_qad.values.reshape(-1,1)))

fin_nn_overtoppingANN_1_transform_qad = pd.DataFrame(scaler.transform(nn_overtoppingANN_1_transform_qad.values.reshape(-1,1)))
print("############## SCORE DATASET TEST FOR ANN ########")
print()
print("R2 TEST")
print(r2_score(ANN_1_q['q'], nn_overtoppingANN_1['Average']))
print()
print("RMSE TEST")
print(np.sqrt(mean_squared_error(ANN_1_q['q'], nn_overtoppingANN_1['Average'])))

train = {
    'features': train_transform,
    'target' : scaler.fit_transform(train_set['qAD'].values.reshape(-1,1))
}
#targetu się nie transformuje
test = {
    'features': test_transform,
    'target' : pd.DataFrame(scaler.fit_transform(test_set['qAD'].values.reshape(-1,1)))
}

test['features'].to_csv('ANN_last.csv', decimal=".", sep='|', index = False ) 


# keep Hm0 data for the inverse_transform
Hm0_test = pd.DataFrame(test_set['Hm0'])


# select the metrics
scoring = ['neg_mean_squared_error','r2']


def transform(Hm0, data,scaler):
    prediction = scaler.inverse_transform(data.values.reshape(-1,1))
    q_overtopping = (10**(prediction.flatten()))*np.sqrt(9.81*(Hm0_test['Hm0'])**3)
    return q_overtopping






#from sklearn.model_selection import GridSearchCV
# RF Regressor

### Add validation data set
# split the train set between train and validation set
train_feature_rf, train_val_feature_rf, train_target_rf, train_val_target_rf = train_test_split(train['features'], train['target'], test_size=0.2,random_state=random_state)
train['features_rf'] = train_feature_rf
train['target_rf'] = train_target_rf
train['val_feature_rf'] = train_val_feature_rf
train['val_target_rf'] = train_val_target_rf


#desc_train_feature_rf = pd.DataFrame(train_feature_rf).describe()
#desc_train_target_rf = pd.DataFrame(train_target_rf).describe()
#desc_train_val_feature_rf = train_val_feature_rf.describe()
#desc_train_val_target_rf = pd.DataFrame(train_val_target_rf).describe()

# build function for hyperparameters search

def RandomForestR(params,  scoring, features, target):
    
    rf = RandomForestRegressor(random_state)
    
    for x in scoring:
        grid_search = GridSearchCV(estimator = rf, param_grid = params, 
                                   n_jobs = -1, verbose = 2, scoring=x)
        
        grid_search.fit(features, target)
        print()
        print("######### Tuning hyper-parameters for %s" % x)
        print()
        print()
        print("Best parameters set found on development set:")
        if x == 'neg_mean_squared_error':
            print(grid_search.best_params_, ' -----> SCORE RMSE: %f' %
                  np.sqrt(-grid_search.best_score_))
        else:
            print(grid_search.best_params_, ' -----> SCORE: %f' %
                  grid_search.best_score_)
# search hyperparameters

scoring = ['neg_mean_squared_error', 'r2']

params_grid = {
    #'bootstrap': [True],
    'max_depth': [ 25],
    #'max_features': ['auto', 'sqrt'],
    'min_samples_leaf': [15, 20, 25],
    #'min_samples_split': [6, 32],
    'n_estimators': [176, 199]
}



RandomForestR(params_grid, scoring, train['features'], train['target'].ravel())
bst_n_estimators =195
#ABS module

rf_best = RandomForestRegressor(max_depth=25, n_estimators=bst_n_estimators, random_state = random_state)
rf_best.fit(train['features_rf'], train['target_rf'])
rf = RandomForestRegressor(     max_depth=25,
                                n_estimators=195,
                                random_state=random_state)
rf.fit(train['features'], train['target'].ravel())

prediction = rf.predict(test['features'])
test['predictions'] = pd.DataFrame(prediction)
print("############## SCORE DATASET TEST FOR rf ########")
print()
print("R2 TEST")
print(r2_score(test['target'], test['predictions']))
print()
print("RMSE TEST")
print(np.sqrt(mean_squared_error(test['target'], test['predictions'])))
print("############## SCORE DATASET TEST FOR ANN ########")
print()
print("R2 TEST")
print(r2_score(fin_ANN_1_qad_test_transform, fin_nn_overtoppingANN_1_transform_qad))
print()
print("RMSE TEST")
print(np.sqrt(mean_squared_error(fin_ANN_1_qad_test_transform, fin_nn_overtoppingANN_1_transform_qad)))

print("############## SCORE DATASET TEST FOR ANN ########")
print()
print("R2 TEST")
print(r2_score(ANN_1_q['q'], nn_overtoppingANN_1['Average']))
print()
print("RMSE TEST")
print(np.sqrt(mean_squared_error(ANN_1_q['q'], nn_overtoppingANN_1['Average'])))
train['features'].to_csv('train_features.csv', decimal=".", sep=' ', header = True, index = False ) 

"""
Created on Sat Dec 31 19:49:44 2022

@author: Dorota Orzechowska
"""
# Import libraries

import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler,MinMaxScaler
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import learning_curve
from sklearn.model_selection import cross_val_score
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error


from sklearn.ensemble import RandomForestRegressor

import seaborn as sns

import warnings
warnings.filterwarnings('ignore')

random_state = 134

# Load the data 
df= pd.read_excel('Database_Eurotop_v5.xlsx')

#dfx= pd.read_excel('Database_Eurotop_v5.xlsx')
df.head()

#create delete list for parameters unused
delete_list = ['Label']

#delete non-core data
df.drop(df.index[df['Core data'] != 'Z'], inplace=True)

#delete rows with Nan values of'q'
df = df.dropna(subset=['q'])

# delete simulation with RF or CF = 4 or 3
df.drop(df.index[df['RF'] == 4], inplace=True)
df.drop(df.index[df['CF'] == 4], inplace=True)
df.drop(df.index[df['RF'] == 3], inplace=True)
df.drop(df.index[df['CF'] == 3], inplace=True)

dfsave = pd.DataFrame({
    'Label':df['Label'],
    'm':df['mmm'],
    'h':df['h'],
    'Hm0 toe':df['Hm0 toe'],
    'Tm1,0t':df['Tm1,0t'],
    'β':df['β'],
    'ht':df['ht'],
    'Bt':df['Bt'],
    'hb':df['hb'],
    'B':df['B'],
    'cotad':df['cotαd'],
    'cotaincl':df['cotαincl'],
    'cotau':df['cotαu'],
    'gf_d':df['gf_d'],
    'gf_u':df['gf_u'],
    'gf':df['gf'],
    'D':df['D'],
    'D50_d':df['D50_d'],
    'D50_u':df['D50_u'],
    'Ac':df['Ac'],
    'Rc':df['Rc'],    
    'Gc':df['Gc'],
    'Kr':df['Kr'],
    'Kt':df['Kt'],
    'q' :df['q']
    })



#delete 'Label' to change data type for float
for x in delete_list:
    dfsave = dfsave.drop(columns=[x])

# convert in float
#dfsave.drop(dfsave.head(1).index, inplace=True)
dfsave = dfsave.astype('float')
dfsave = dfsave.dropna(subset=['m'])
dfsave = dfsave.dropna(subset=['D'])
a = dfsave.describe()



#delete q == 0
dfsave.drop(dfsave.index[dfsave['q'] == 0 ], inplace=True)


#%matplotlib qt

for i in dfsave.columns:
    fig = plt.hist(dfsave[i],bins=10)
    plt.title(i)
    plt.show(fig)

# delete the NaN 
#dfsave = dfsave.dropna()
#a = df.describe()

#corr matrix to see correlation of parameters
corr_matrix = dfsave.corr()
f, ax = plt.subplots(figsize=(20,20))
g = sns.heatmap(corr_matrix, cmap = sns.diverging_palette(10, 250, as_cmap = True), annot = True)

#creating an additional feature - compute spectral wave length in deep water lm0 in meter
dfsave['Lm0'] = ((9.81*(dfsave['Tm1,0t'])**2)/(2*np.pi))

# create dataframe containing the model parameters
dff = pd.DataFrame({
    'Lm0' :dfsave['Lm0'],
    'Hm0' :dfsave['Hm0 toe'],
    'Hm0/Lm0':dfsave['Hm0 toe']/dfsave['Lm0'],
    'β':dfsave['β'],
    'h/Lm0':dfsave['h']/dfsave['Lm0'],
    'ht/Hm0':dfsave['ht']/dfsave['Hm0 toe'],
    'Bt/Lm0':dfsave['Bt']/dfsave['Lm0'],
    'hb/Hm0':dfsave['hb']/dfsave['Hm0 toe'],
    'B/Lm0':dfsave['B']/dfsave['Lm0'],
    'Ac/Hm0':dfsave['Ac']/dfsave['Hm0 toe'],
    'Rc/Hm0':dfsave['Rc']/dfsave['Hm0 toe'],
    #'Rc':dfsave['Rc'],
    'Gc/Lm0':dfsave['Gc']/dfsave['Lm0'],
    'm':dfsave['m'],
    'cotad':dfsave['cotad'],
    'cotaincl':dfsave['cotaincl'],
    'gf':dfsave['gf'],
    'q':dfsave['q'],
    'D/Hm0':dfsave['D']/dfsave['Hm0 toe'],
    #target:
    'qAD':np.log10(dfsave['q']/np.sqrt(9.81*(dfsave['Hm0 toe'])**3)),
    })
#dff.head()

dff.describe()

#dff = dff.dropna()

b= dff.describe()

# Replace infinite updated data with nan
dff.replace([np.inf, -np.inf], np.nan, inplace=True)
# Drop rows with NaN
dff.dropna(inplace=True)

#dfsave.to_csv('NN_OVERTOPPING6.csv', decimal=".", sep=' ', header = False, index = False ) 
#val_=dfsave.describe()

corr_matrix = dff.corr()
f, ax = plt.subplots(figsize=(20,20))
g = sns.heatmap(corr_matrix, cmap = sns.diverging_palette(10, 250, as_cmap = True), annot = True)


train_set, test_set = train_test_split(dff, test_size=0.2, random_state = random_state)

#scaler_std = StandardScaler()
scaler = StandardScaler()

ANN_1 = dfsave.loc[test_set.index] 
ANN_1_q = pd.DataFrame(ANN_1['q'])
ANN_1_qad = pd.DataFrame(np.log10(ANN_1['q']/np.sqrt(9.81*(ANN_1['Hm0 toe'])**3)))

ANN_1 = ANN_1.drop(['Lm0', 'D', 'gf','Kr', 'Kt','q', 'cotaincl'], axis = 1)

ANN_1['Kr'] = 0
ANN_1['Kt'] = 0
ANN_1['q'] = 1
ANN_1.insert(loc=0, column='Label', value=df['Label'])
ANN_1.to_csv('ANN_1.csv', decimal=".", sep='|', index = False )
nn_overtoppingANN_1 = pd.read_csv('ANN_results.csv', decimal=".", sep='|')

nn_overtoppingANN_1_transform = pd.DataFrame(nn_overtoppingANN_1['Average'])

ANN_1_Hm = pd.DataFrame(ANN_1["Hm0 toe"])
ANN_1_Hm=ANN_1_Hm.reset_index()
ANN_1_Hm = ANN_1_Hm.drop(['index'], axis=1)

nn_overtoppingANN_1_transform_qad = pd.DataFrame(np.log10(nn_overtoppingANN_1_transform['Average']/np.sqrt(9.81*(ANN_1_Hm['Hm0 toe'])**3)))


# save description of train data
train_transform = pd.DataFrame(train_set.drop(['Lm0', 'Hm0','q','qAD'], axis=1))
desc_train_transform = train_transform.describe()
desc_train_transform.to_excel('desc_train_transform.xlsx') 
train_transform = pd.DataFrame(scaler.fit_transform(train_transform))
train_transform.columns = train_set.drop(['Lm0','Hm0','q','qAD'], axis=1).columns

test_transform = pd.DataFrame(test_set.drop(['Lm0','Hm0','q','qAD'], axis=1))

#save description of test data
desc_test_transform = test_transform.describe()
desc_test_transform.to_csv('desc_test_transform.csv', decimal=".", sep=' ', header = True, index = False ) 

#add noise i zapisz ponownie
#jesli dla zaszumionego datasetu będzie podobny wynik albo mniejszy - to jest ok, a jesli wiekszy to overfitting i 
test_transform = pd.DataFrame(scaler.transform(test_transform))
test_transform.columns = test_set.drop(['Lm0','Hm0','q','qAD'], axis=1).columns

fin_ANN_1_qad_test_transform = pd.DataFrame(scaler.fit_transform(ANN_1_qad.values.reshape(-1,1)))

fin_nn_overtoppingANN_1_transform_qad = pd.DataFrame(scaler.transform(nn_overtoppingANN_1_transform_qad.values.reshape(-1,1)))
test['features'].to_csv('ANN_last.csv', decimal=".", sep='|', index = False ) 

train['features'].to_csv('train_features.csv', decimal=".", sep=' ', header = True, index = False ) 

# keep Hm0 data for the inverse_transform
Hm0_test = pd.DataFrame(test_set['Hm0'])


# select the metrics
scoring = ['neg_mean_squared_error','r2']


def transform(Hm0, data,scaler):
    prediction = scaler.inverse_transform(data.values.reshape(-1,1))
    q_overtopping = (10**(prediction.flatten()))*np.sqrt(9.81*(Hm0_test['Hm0'])**3)
    return q_overtopping
train_feature_rf, train_val_feature_rf, train_target_rf, train_val_target_rf = train_test_split(train['features'], train['target'], test_size=0.2,random_state=random_state)
train['features_rf'] = train_feature_rf
train['target_rf'] = train_target_rf
train['val_feature_rf'] = train_val_feature_rf
train['val_target_rf'] = train_val_target_rf


#desc_train_feature_rf = pd.DataFrame(train_feature_rf).describe()
#desc_train_target_rf = pd.DataFrame(train_target_rf).describe()
#desc_train_val_feature_rf = train_val_feature_rf.describe()
#desc_train_val_target_rf = pd.DataFrame(train_val_target_rf).describe()

# build function for hyperparameters search

def RandomForestR(params,  scoring, features, target):
    
    rf = RandomForestRegressor(random_state)
    
    for x in scoring:
        grid_search = GridSearchCV(estimator = rf, param_grid = params, 
                                   n_jobs = -1, verbose = 2, scoring=x)
        
        grid_search.fit(features, target)
        print()
        print("######### Tuning hyper-parameters for %s" % x)
        print()
        print()
        print("Best parameters set found on development set:")
        if x == 'neg_mean_squared_error':
            print(grid_search.best_params_, ' -----> SCORE RMSE: %f' %
                  np.sqrt(-grid_search.best_score_))
        else:
            print(grid_search.best_params_, ' -----> SCORE: %f' %
                  grid_search.best_score_)
# search hyperparameters

scoring = ['neg_mean_squared_error', 'r2']


params_grid = {
    #'bootstrap': [True],
    'max_depth': [ 25],
    #'max_features': ['auto', 'sqrt'],
    'min_samples_leaf': [15, 20, 25],
    #'min_samples_split': [6, 32],
    'n_estimators': [176, 199]
}
bst_n_estimators =195
#ABS module

rf_best = RandomForestRegressor(max_depth=25, n_estimators=bst_n_estimators, random_state = random_state)
rf_best.fit(train['features_rf'], train['target_rf'])
def features_importances(estimator, features):
    importances = estimator.feature_importances_
    indices = np.argsort(importances)[::-1]
    plt.figure(figsize=(15, 10))
    plt.title("Feature importances")
    plt.bar(range(len(importances)),
            importances[indices], color="blue", align="center")
    plt.xticks(range(len(importances)),
               features.columns[indices], fontsize=10, rotation=90)
    plt.xlim([-1, len(importances)])
    plt.show()

features_importances(rf,train['features'])
rf = RandomForestRegressor(     max_depth=25,
                                n_estimators=195,
                                random_state=random_state)
rf.fit(train['features'], train['target'].ravel())

prediction = rf.predict(test['features'])
test['predictions'] = pd.DataFrame(prediction)
print("############## SCORE DATASET TEST FOR rf ########")
print()
print("R2 TEST")
print(r2_score(test['target'], test['predictions']))
print()
print("RMSE TEST")
print(np.sqrt(mean_squared_error(test['target'], test['predictions'])))
print("############## SCORE DATASET TEST FOR ANN ########")
print()
print("R2 TEST")
print(r2_score(fin_ANN_1_qad_test_transform, fin_nn_overtoppingANN_1_transform_qad))
print()
print("RMSE TEST")
print(np.sqrt(mean_squared_error(fin_ANN_1_qad_test_transform, fin_nn_overtoppingANN_1_transform_qad)))

## ---(Tue Apr 11 15:58:22 2023)---
import openpyxl

# Wczytanie danych z arkusza kalkulacyjnego
wb = openpyxl.load_workbook('nazwa_pliku.xlsx')
sheet = wb.active
n = int(sheet.cell(row=1, column=1).value)
m = int(sheet.cell(row=1, column=2).value)
kaczuszki = []
for i in range(n):
    w = int(sheet.cell(row=i+2, column=1).value)
    s = int(sheet.cell(row=i+2, column=2).value)
    kaczuszki.append((w, s))

# Algorytm szukający optymalnego rozwiązania
dp = [[0 for j in range(m+1)] for i in range(n+1)]
for i in range(1, n+1):
    for j in range(1, m+1):
        dp[i][j] = dp[i-1][j]
        if j >= kaczuszki[i-1][1]:
            dp[i][j] = max(dp[i][j], dp[i-1][j-kaczuszki[i-1][1]]+kaczuszki[i-1][0])

# Wyświetlenie wyniku
print(dp[n][m])
import openpyxl

wb = openpyxl.load_workbook('zadanie-rekrutacyjne.xlsx')
sheet = wb.active
n = int(sheet.cell(row=1, column=1).value)
m = int(sheet.cell(row=1, column=2).value)
kaczuszki = []
for i in range(n):
    w = int(sheet.cell(row=i+2, column=1).value)
    s = int(sheet.cell(row=i+2, column=2).value)
    kaczuszki.append((w, s))

# Algorytm szukający optymalnego rozwiązania
dp = [[0 for j in range(m+1)] for i in range(n+1)]
for i in range(1, n+1):
    for j in range(1, m+1):
        dp[i][j] = dp[i-1][j]
        if j >= kaczuszki[i-1][1]:
            dp[i][j] = max(dp[i][j], dp[i-1][j-kaczuszki[i-1][1]]+kaczuszki[i-1][0])

# Wyświetlenie wyniku
print(dp[n][m])
import openpyxl

# wczytanie danych z arkusza kalkulacyjnego
arkusz = openpyxl.load_workbook('nazwa_pliku.xlsx')
arkusz_roboczy = arkusz.active
N = arkusz_roboczy.cell(row=1, column=1).value
M = arkusz_roboczy.cell(row=1, column=2).value
kaczuszki = []
for i in range(2, N+2):
    wysokosc = arkusz_roboczy.cell(row=i, column=1).value
    szerokosc = arkusz_roboczy.cell(row=i, column=2).value
    kaczuszki.append((wysokosc, szerokosc))

# sortowanie kaczuszek wg malejącej wysokości
kaczuszki.sort(reverse=True)

# ustawienie kaczuszek w rzędzie, zaczynając od lewej strony
suma_wysokosci = 0
szerokosc_rzedu = 0
for k in kaczuszki:
    if szerokosc_rzedu + k[1] <= M:
        suma_wysokosci += k[0]
        szerokosc_rzedu += k[1]
    else:
        break

# wyświetlenie wyniku
print(suma_wysokosci)
import openpyxl

# wczytanie danych z arkusza kalkulacyjnego
arkusz = openpyxl.load_workbook('zadanie-rekrutacyjne.xlsx')
arkusz_roboczy = arkusz.active
N = arkusz_roboczy.cell(row=1, column=1).value
M = arkusz_roboczy.cell(row=1, column=2).value
kaczuszki = []
for i in range(2, N+2):
    wysokosc = arkusz_roboczy.cell(row=i, column=1).value
    szerokosc = arkusz_roboczy.cell(row=i, column=2).value
    kaczuszki.append((wysokosc, szerokosc))

# sortowanie kaczuszek wg malejącej wysokości
kaczuszki.sort(reverse=True)

# ustawienie kaczuszek w rzędzie, zaczynając od lewej strony
suma_wysokosci = 0
szerokosc_rzedu = 0
for k in kaczuszki:
    if szerokosc_rzedu + k[1] <= M:
        suma_wysokosci += k[0]
        szerokosc_rzedu += k[1]
    else:
        break

# wyświetlenie wyniku
print(suma_wysokosci)

## ---(Tue May  9 17:11:50 2023)---
runfile('C:/Users/to ja/.spyder-py3/vibration.py', wdir='C:/Users/to ja/.spyder-py3')

## ---(Tue May  9 17:16:02 2023)---
import pygame
import sys
from pygame.locals import *
import numpy as np
from numpy.linalg import inv
from spring import spring

## ---(Sat May 20 13:46:48 2023)---
runfile('C:/Users/to ja/.spyder-py3/untitled0.py', wdir='C:/Users/to ja/.spyder-py3')
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from bs4.element import Tag
from time import sleep
import csv
from parsel import Selector
import parameters
import numpy

# Function call extracting title and linkedin profile iteratively
def find_profiles():
    for r in result_div:
        # Checks if each element is present, else, raise exception
        try:
            link = r.find('a', href=True)
            title = None
            title = r.find('h3')
            
            # returns True if a specified object is of a specified type; Tag in this instance 
            if isinstance(title,Tag):
                title = title.get_text()
            
            description = None
            description = r.find('span', attrs={'class': 'st'})
            
            if isinstance(description, Tag):
                description = description.get_text()
            
            # Check to make sure everything is present before appending
            if link != '' and title != '' and description != '':
                links.append(link['href'])
                titles.append(title)
                descriptions.append(description)
        
        
        # Next loop if one element is not present
        except Exception as e:
            print(e)
            continue

# This function iteratively clicks on the "Next" button at the bottom right of the search page. 
def profiles_loop():
    
    find_profiles()
    
    next_button = driver.find_element_by_xpath('//*[@id="pnnext"]') 
    next_button.click()


def repeat_fun(times, f):
    for i in range(times): f()

# specifies the path to the chromedriver.exe
driver = webdriver.Chrome(r'C:\Users\to ja\Downloads\chromedriver_win32')


# driver.get method() will navigate to a page given by the URL address
driver.get('https://www.linkedin.com')

# locate email form by_class_name
username = driver.find_element_by_id('session_key')

# send_keys() to simulate key strokes
username.send_keys("u1912475@gmail.com")
sleep(0.5)

# locate password form by_class_name
password = driver.find_element_by_id('session_password')

# send_keys() to simulate key strokes
password.send_keys("5742191u")
sleep(0.5)

# locate submit button by_class_name
log_in_button = driver.find_element_by_class_name('sign-in-form__submit-button')

# .click() to mimic button click
log_in_button.click()
sleep(0.5)
runfile('C:/Users/to ja/.spyder-py3/untitled0.py', wdir='C:/Users/to ja/.spyder-py3')
runcell(0, 'C:/Users/to ja/.spyder-py3/untitled0.py')
driver.get('https://www.linkedin.com/home')
runcell(0, 'C:/Users/to ja/.spyder-py3/untitled0.py')
runfile('C:/Users/to ja/.spyder-py3/untitled0.py', wdir='C:/Users/to ja/.spyder-py3')
runcell(0, 'C:/Users/to ja/.spyder-py3/untitled0.py')
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from bs4.element import Tag
from time import sleep
import csv
from parsel import Selector
import parameters
import numpy



# specifies the path to the chromedriver.exe
driver = webdriver.Chrome(r"C:\Users\to ja\.spyder-py3\chromedriver_win32.exe")


# driver.get method() will navigate to a page given by the URL address
driver.get('https://www.linkedin.com/home')

# locate email form by_class_name
username = driver.find_element_by_xpath("//input[@name = 'session_key']")

# send_keys() to simulate key strokes
username.send_keys("u1912475@gmail.com")


# locate password form by_class_name
password = driver.find_element_by_xpath("//input[@name = 'session_password']")

# send_keys() to simulate key strokes
password.send_keys("5742191u")


# locate submit button by_class_name
log_in_button = driver.find_element_by_class_name('sign-in-form__submit-button')

# .click() to mimic button click
log_in_button.click()
sleep(0.5)
runfile('C:/Users/to ja/.spyder-py3/untitled0.py', wdir='C:/Users/to ja/.spyder-py3')